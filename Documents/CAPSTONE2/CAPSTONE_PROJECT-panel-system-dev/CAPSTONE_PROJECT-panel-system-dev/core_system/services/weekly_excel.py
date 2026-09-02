from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from django.utils import timezone

from core_system.models import (
    TransactionArchive,
    Contribution,
    AidTrackingPost,
)


def _style_header(ws, row, cols):
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    header_fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
    header_font = Font(color="ffffff", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _auto_width(ws, cols):
    for col in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[chr(64 + col)].width = max_len + 4


def _fmt(val):
    if val is None:
        return 0.0
    return round(float(val), 2)


def _month_range(year, month):
    from datetime import date, timedelta
    start = date(year, month, 1)
    if month == 12:
        end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(year, month + 1, 1) - timedelta(days=1)
    return start, end


def build_cash_flow_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14, color="1b5e20")

    today = timezone.localdate()
    current_year = today.year
    current_month = today.month

    prior_year = current_year if current_month > 1 else current_year - 1
    prior_month = current_month - 1 if current_month > 1 else 12

    prior_start, prior_end = _month_range(prior_year, prior_month)
    current_start, _ = _month_range(current_year, current_month)

    def archived_in_range(start_date, end_date, tx_types):
        return TransactionArchive.objects.filter(
            transaction_type__in=tx_types,
            archived_at__date__gte=start_date,
            archived_at__date__lte=end_date,
        ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

    def contributions_paid_in_range(start_date, end_date):
        return Contribution.objects.filter(
            status__in=["PAID", "RECORDED", "PENDING_VERIFICATION"],
            payment_date__gte=start_date,
            payment_date__lte=end_date,
        ).aggregate(total=Sum("paid_amount"))["total"] or Decimal("0.00")

    def pending_contributions():
        return Contribution.objects.filter(
            status="NOT_PAID",
            aid_tracking_post_id_FK__is_active=True,
        ).aggregate(total=Sum("expected_amount"))["total"] or Decimal("0.00")

    prior_dues = archived_in_range(prior_start, prior_end, ["monthly_dues"])
    prior_fees = archived_in_range(prior_start, prior_end, ["membership_fee"])
    prior_contribs = contributions_paid_in_range(prior_start, prior_end)
    prior_medical = archived_in_range(prior_start, prior_end, ["medical_aid"])
    prior_death = archived_in_range(prior_start, prior_end, ["death_aid"])

    curr_dues = archived_in_range(current_start, today, ["monthly_dues"])
    curr_fees = archived_in_range(current_start, today, ["membership_fee"])
    curr_contribs = contributions_paid_in_range(current_start, today)
    curr_medical = archived_in_range(current_start, today, ["medical_aid"])
    curr_death = archived_in_range(current_start, today, ["death_aid"])

    pending = pending_contributions()

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Statement"

    ws.cell(row=1, column=1, value="CAUFA — Cash Flow Statement").font = title_font
    ws.merge_cells("A1:C1")

    sub_headers = ["Category", "Debit (₱)", "Credit (₱)"]
    for ci, h in enumerate(sub_headers, 1):
        ws.cell(row=3, column=ci, value=h)
    _style_header(ws, 3, 3)

    row = 4

    def section_row(label):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="eef7ef", end_color="eef7ef", fill_type="solid"
        )
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    def data_row(label, debit, credit):
        nonlocal row
        ws.cell(row=row, column=1, value=label).border = thin_border
        ws.cell(row=row, column=2, value=_fmt(debit)).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=_fmt(credit)).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
        row += 1

    def total_row(label, debit, credit):
        nonlocal row
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).font = bold_font
        ws.cell(row=row, column=1, value=label)
        net = _fmt(debit - credit)
        ws.cell(row=row, column=2, value=_fmt(debit))
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=3, value=_fmt(credit))
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")
        row += 1
        ws.cell(row=row, column=1, value="Net Cash Flow")
        ws.cell(row=row, column=1).font = bold_font
        ws.cell(row=row, column=2, value=_fmt(net))
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=2).font = bold_font
        if net >= 0:
            ws.cell(row=row, column=2).fill = green_fill
        else:
            ws.cell(row=row, column=2).fill = red_fill
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    prior_label = f"Prior Month ({prior_year}-{prior_month:02d})"
    section_row(prior_label)
    data_row("Monthly Dues", prior_dues, Decimal("0.00"))
    data_row("Membership Fees", prior_fees, Decimal("0.00"))
    data_row("Aid Contributions Collected", prior_contribs, Decimal("0.00"))
    data_row("Medical Aid Releases", Decimal("0.00"), prior_medical)
    data_row("Death Aid Releases", Decimal("0.00"), prior_death)
    prior_debit_total = prior_dues + prior_fees + prior_contribs
    prior_credit_total = prior_medical + prior_death
    total_row("Subtotal", prior_debit_total, prior_credit_total)

    curr_label = f"Current Month ({current_year}-{current_month:02d}) to date"
    section_row(curr_label)
    data_row("Monthly Dues", curr_dues, Decimal("0.00"))
    data_row("Membership Fees", curr_fees, Decimal("0.00"))
    data_row("Aid Contributions Collected", curr_contribs, Decimal("0.00"))
    data_row("Medical Aid Releases", Decimal("0.00"), curr_medical)
    data_row("Death Aid Releases", Decimal("0.00"), curr_death)
    curr_debit_total = curr_dues + curr_fees + curr_contribs
    curr_credit_total = curr_medical + curr_death
    total_row("Subtotal", curr_debit_total, curr_credit_total)

    section_row("Pending (Expected but not yet Collected)")
    data_row("Unpaid Aid Contributions", pending, Decimal("0.00"))
    data_row("", Decimal("0.00"), Decimal("0.00"))
    total_debit = prior_debit_total + curr_debit_total
    total_credit = prior_credit_total + curr_credit_total
    total_row("Grand Total", total_debit, total_credit)

    _auto_width(ws, 3)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
