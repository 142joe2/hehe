"""Oversight report builders for the President Dashboard.

Each builder takes a request and returns a normalized ``report`` dict that the
ReportPreview frontend can render uniformly, plus a ``legacy`` dict kept for
backward compatibility with the existing render functions in
``oversight_reports.js``.

Normalized report schema
------------------------
{
    "report_key": str,
    "report_name": str,
    "description": str,
    "generated_by": str,
    "generated_at": str (iso "YYYY-MM-DD HH:MM:SS"),
    "filters": [{"label": str, "value": str}],
    "summary": [{"label": str, "value": number|str, "type": "count"|"currency"|"percent"|"string"}],
    "columns": [{"key": str, "label": str, "align": "left"|"right"|"center"}],
    "rows": [ {key: value}, ... ],
    "sections": [{"title": str, "columns": [...], "rows": [...]}],   # optional
    "notes": [str],                                                  # optional
}
"""

from __future__ import annotations

import base64
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.db.models import Count, Q, Sum
from django.utils import timezone

from core_system.constants.policy_constants import get_contribution_amount_for_aid, get_monthly_dues_amount
from core_system.constants.status_constants import Status
from core_system.models import (
    AidTrackingPost,
    Contribution,
    DeathAid,
    FundTransaction,
    GlobalAuditTrail,
    MedicalAid,
    Member,
    MonthlyDues,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

AID_TYPE_LABELS = {
    "medical_aid": "Medical Aid",
    "death_aid": "Death Aid",
    "Medical Aid": "Medical Aid",
    "Death Aid": "Death Aid",
}

DEATH_CATEGORY_LABELS = {
    "member": "Member",
    "spouse": "Husband/Wife",
    "parent_child": "Parents and Children",
    "sibling": "Brother/Sister (Full Blood)",
    "other": "Other",
}

FINAL_CLAIM_STATUSES = set(Status.ALL_FINAL) | {"Completed", "Complete"}

PAID_STATUSES = set(Status.ALL_AUDITOR_VERIFIED)
PENDING_STATUSES = set(Status.ALL_PENDING)

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def _now_str() -> str:
    return timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M:%S")


def _get(request, key, default=""):
    return (request.GET.get(key) or "").strip() or default


def _active_members():
    return Member.objects.exclude(membership_status__iexact="retired")


def _apply_college(qs, college, member_related="member_id_FK"):
    if college:
        return qs.filter(**{f"{member_related}__department__iexact": college})
    return qs


def _report_base(report_key, report_name, description, request, filters):
    session = getattr(request, "session", None)
    generated_by = getattr(request, "officer_name", None) or (session.get("officer_name") if session else None) or "President"
    return {
        "report_key": report_key,
        "report_name": report_name,
        "description": description,
        "generated_by": generated_by,
        "generated_at": _now_str(),
        "filters": filters,
    }


def _col(key, label, align="left"):
    return {"key": key, "label": label, "align": align}


def _as_float(value) -> float:
    return float(value or 0)


def _latest_approval_datetime(table_name, record_id):
    """Return the latest approval action timestamp for a claim record."""
    trail = (
        GlobalAuditTrail.objects.filter(
            table_name__iexact=table_name,
            record_id=str(record_id),
            action__icontains="approved",
        )
        .order_by("-timestamp")
        .first()
    )
    if trail and trail.timestamp:
        return trail.timestamp.strftime("%Y-%m-%d")
    return "N/A"


def _release_info_by_post():
    """Map AidTrackingPost id -> release date from audit trail (FINISH_RELEASED / RELEASE_NOTIFIED)."""
    result = {}
    for trail in (
        GlobalAuditTrail.objects.filter(
            table_name__iexact="AID_TRACKING_POST",
            action__in=["FINISH_RELEASED", "RELEASE_NOTIFIED"],
        )
        .order_by("-timestamp")
    ):
        if trail.record_id not in result and trail.timestamp:
            result[trail.record_id] = trail.timestamp.strftime("%Y-%m-%d")
    return result


def _contribution_totals_for_source(source_type, source_id):
    """Expected + collected amounts for a claim, from contributions of its tracking post."""
    post = AidTrackingPost.objects.filter(source_type=source_type, source_id=str(source_id)).first()
    if post is None:
        return 0.0, 0.0
    expected = post.total_expected or 0
    collected = post.total_collected or 0
    if not expected:
        agg = Contribution.objects.filter(aid_tracking_post_id_FK=post).aggregate(
            exp=Sum("expected_amount"), col=Sum("paid_amount")
        )
        expected = agg["exp"] or 0
        collected = agg["col"] or 0
    return _as_float(expected), _as_float(collected)


# ---------------------------------------------------------------------------
# 1. Members by College
# ---------------------------------------------------------------------------

def build_members_by_college(request):
    college = _get(request, "college")
    membership_status = _get(request, "membership_status")
    payment_status = _get(request, "payment_status")
    year = _get(request, "year")
    month = _get(request, "month")

    members = _active_members()
    if college:
        members = members.filter(department__iexact=college)
    if membership_status:
        members = members.filter(membership_status__iexact=membership_status)

    month_key = ""
    paid_ids, pending_ids = set(), set()
    if year and month:
        month_key = f"{year}-{month.zfill(2)}"
        dues = MonthlyDues.objects.filter(month_covered=month_key)
        paid_ids = set(dues.filter(payment_status__in=list(PAID_STATUSES)).values_list("member_id_FK_id", flat=True))
        pending_ids = set(dues.filter(payment_status__in=list(PENDING_STATUSES)).values_list("member_id_FK_id", flat=True))

    if payment_status == "paid":
        members = members.filter(member_id_PK__in=paid_ids)
    elif payment_status == "unpaid":
        members = members.exclude(member_id_PK__in=paid_ids)
    elif payment_status == "pending":
        members = members.filter(member_id_PK__in=pending_ids)

    grouped: "OrderedDict[str, dict]" = OrderedDict()
    for member in members.order_by("department", "full_name"):
        dept = member.department or "Unassigned"
        entry = grouped.setdefault(
            dept,
            {
                "college": dept,
                "total_members": 0,
                "active_members": 0,
                "paid_members": 0,
                "unpaid_members": 0,
                "members": [],
            },
        )
        entry["total_members"] += 1
        if member.membership_status and member.membership_status.lower() != "retired":
            entry["active_members"] += 1

        payment_info = "N/A"
        if month_key:
            due = MonthlyDues.objects.filter(
                member_id_FK_id=member.member_id_PK, month_covered=month_key
            ).first()
            payment_info = due.payment_status if due else "Unpaid"
            if payment_info in PAID_STATUSES:
                entry["paid_members"] += 1
            elif payment_info not in PENDING_STATUSES:
                entry["unpaid_members"] += 1

        entry["members"].append({
            "member_id": member.member_id_PK,
            "full_name": member.full_name,
            "position": member.position or "N/A",
            "college": dept,
            "department": dept,
            "membership_status": member.membership_status or "Unknown",
            "payment_status": payment_info,
        })

    report_data = list(grouped.values())
    total_members = sum(e["total_members"] for e in report_data)

    filters = [
        {"label": "College", "value": college or "All"},
        {"label": "Membership Status", "value": membership_status or "All"},
        {"label": "Payment Status", "value": payment_status or "All"},
        {"label": "Period", "value": f"{MONTH_NAMES[int(month) - 1] if month.isdigit() and 1 <= int(month) <= 12 else month or 'All'} {year or ''}".strip() or "All"},
    ]

    report = _report_base(
        "members_by_college", "Members by College",
        "Distribution of active members grouped by college with payment status for the selected period.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Total Colleges", "value": len(report_data), "type": "count"},
        {"label": "Total Members", "value": total_members, "type": "count"},
    ]
    report["columns"] = [
        _col("college", "College", "left"),
        _col("total_members", "Total Members", "right"),
        _col("active_members", "Active", "right"),
        _col("paid_members", "Paid", "right"),
        _col("unpaid_members", "Unpaid", "right"),
    ]
    report["rows"] = [
        {
            "college": e["college"],
            "total_members": e["total_members"],
            "active_members": e["active_members"],
            "paid_members": e["paid_members"],
            "unpaid_members": e["unpaid_members"],
        }
        for e in report_data
    ]

    legacy = {
        "filters": {
            "college": college,
            "membership_status": membership_status,
            "payment_status": payment_status,
            "year": year,
            "month": month,
        },
        "report_data": report_data,
        "summary": {
            "total_colleges": len(report_data),
            "total_members": total_members,
            "generated_at": _now_str(),
        },
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 2. Paid / Unpaid Summary
# ---------------------------------------------------------------------------

def build_paid_unpaid_summary(request):
    year = _get(request, "year") or str(timezone.now().year)
    month = _get(request, "month") or str(timezone.now().month).zfill(2)
    college = _get(request, "college")

    month_key = f"{year}-{month.zfill(2)}"
    members = _active_members()
    if college:
        members = members.filter(department__iexact=college)

    total_members = members.count()

    dues = MonthlyDues.objects.filter(month_covered=month_key)
    if college:
        dues = dues.filter(member_id_FK__department__iexact=college)

    paid_count = dues.filter(payment_status__in=list(PAID_STATUSES)).count()
    pending_count = dues.filter(payment_status__in=list(PENDING_STATUSES)).count()
    unpaid_count = max(total_members - paid_count - pending_count, 0)
    total_collected = dues.filter(payment_status__in=list(PAID_STATUSES)).aggregate(total=Sum("amount"))["total"] or 0

    monthly_dues_amount = get_monthly_dues_amount()
    expected_collection = (total_members * monthly_dues_amount) if monthly_dues_amount else 0

    def pct(n):
        return round((n / total_members * 100), 2) if total_members else 0

    department_breakdown = []
    for dept in members.values_list("department", flat=True).distinct().order_by("department"):
        if not dept:
            continue
        dept_total = members.filter(department=dept).count()
        dept_dues = dues.filter(member_id_FK__department=dept)
        dept_paid = dept_dues.filter(payment_status__in=list(PAID_STATUSES)).count()
        dept_pending = dept_dues.filter(payment_status__in=list(PENDING_STATUSES)).count()
        dept_unpaid = max(dept_total - dept_paid - dept_pending, 0)
        department_breakdown.append({
            "college": dept,
            "total": dept_total,
            "paid": dept_paid,
            "pending": dept_pending,
            "unpaid": dept_unpaid,
            "compliance": round((dept_paid / dept_total * 100) if dept_total else 0, 2),
        })

    filters = [
        {"label": "Period", "value": f"{MONTH_NAMES[int(month) - 1]} {year}"},
        {"label": "College", "value": college or "All"},
    ]

    report = _report_base(
        "paid_unpaid_summary", "Paid / Unpaid Summary",
        "Monthly dues collection summary showing paid, unpaid and pending members by college.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Total Members", "value": total_members, "type": "count"},
        {"label": "Paid", "value": paid_count, "type": "count"},
        {"label": "Unpaid", "value": unpaid_count, "type": "count"},
        {"label": "Pending", "value": pending_count, "type": "count"},
        {"label": "Paid Percentage", "value": pct(paid_count), "type": "percent"},
        {"label": "Total Collected", "value": _as_float(total_collected), "type": "currency"},
        {"label": "Expected Collection", "value": _as_float(expected_collection), "type": "currency"},
        {"label": "Collection Rate", "value": round((_as_float(total_collected) / _as_float(expected_collection) * 100) if _as_float(expected_collection) else 0, 2), "type": "percent"},
    ]
    report["columns"] = [
        _col("college", "College", "left"),
        _col("total", "Total", "right"),
        _col("paid", "Paid", "right"),
        _col("pending", "Pending", "right"),
        _col("unpaid", "Unpaid", "right"),
        _col("compliance", "Compliance", "right"),
    ]
    report["rows"] = department_breakdown

    legacy = {
        "filters": {"year": year, "month": month, "college": college},
        "summary": {
            "total_members": total_members,
            "paid_members": paid_count,
            "unpaid_members": unpaid_count,
            "pending_members": pending_count,
            "paid_percentage": pct(paid_count),
            "unpaid_percentage": pct(unpaid_count),
            "pending_percentage": pct(pending_count),
            "total_collected": _as_float(total_collected),
            "expected_collection": _as_float(expected_collection),
            "collection_rate": round((_as_float(total_collected) / _as_float(expected_collection) * 100) if _as_float(expected_collection) else 0, 2),
        },
        "department_breakdown": department_breakdown,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 3. Pending Claims
# ---------------------------------------------------------------------------

STAGE_MAP = {
    "Pending": "Treasurer Review",
    "Pending Verification": "Treasurer Review",
    "Pending Treasurer Check": "Treasurer Review",
    "Treasurer Direct": "Treasurer Review",
    "Pending Auditor Verification": "Auditor Verification",
    "Pending Auditor Review": "Auditor Verification",
    "Under Review": "Auditor Verification",
    "Auditor Verified": "President Approval",
    "Approved": "President Approval",
    "Verified": "Contribution Collection",
    "Released": "Released",
    "Rejected": "Rejected",
}


def build_pending_claims(request):
    claim_type = _get(request, "claim_type")
    college = _get(request, "college")

    pending_statuses = list(PENDING_STATUSES | {"Under Review", "Auditor Verified", "Verified"})
    medical_claims, death_claims = [], []

    if claim_type in ("", "medical"):
        qs = MedicalAid.objects.filter(status__in=pending_statuses)
        qs = _apply_college(qs, college)
        for claim in qs:
            medical_claims.append({
                "claim_id": claim.medical_aid_id_PK,
                "member_id": claim.member_id_FK.member_id_PK,
                "member_name": claim.member_id_FK.full_name,
                "college": claim.member_id_FK.department or "N/A",
                "claim_type": "Medical Aid",
                "diagnosis": claim.reason_for_request or "N/A",
                "amount_requested": _as_float(claim.requested_amount),
                "date_filed": claim.request_date.strftime("%Y-%m-%d") if claim.request_date else "N/A",
                "status": claim.status,
                "current_stage": STAGE_MAP.get(claim.status, "Unknown"),
            })

    if claim_type in ("", "death"):
        qs = DeathAid.objects.filter(status__in=pending_statuses)
        qs = _apply_college(qs, college)
        for claim in qs:
            death_claims.append({
                "claim_id": claim.death_aid_id_PK,
                "member_id": claim.member_id_FK.member_id_PK,
                "member_name": claim.member_id_FK.full_name,
                "college": claim.member_id_FK.department or "N/A",
                "claim_type": "Death Aid",
                "deceased_name": claim.deceased_name or "N/A",
                "relationship": claim.relationship_to_member or "N/A",
                "amount_requested": _as_float(claim.benefit_amount),
                "date_filed": claim.claim_date.strftime("%Y-%m-%d") if claim.claim_date else "N/A",
                "status": claim.status,
                "current_stage": STAGE_MAP.get(claim.status, "Unknown"),
            })

    all_claims = medical_claims + death_claims
    total_amount = sum(c["amount_requested"] for c in all_claims)

    filters = [
        {"label": "Claim Type", "value": {"medical": "Medical Aid", "death": "Death Aid"}.get(claim_type, "All")},
        {"label": "College", "value": college or "All"},
    ]

    report = _report_base(
        "pending_claims", "Pending Claims",
        "Claims currently awaiting Treasurer, Auditor or President action.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Total Claims", "value": len(all_claims), "type": "count"},
        {"label": "Medical Claims", "value": len(medical_claims), "type": "count"},
        {"label": "Death Claims", "value": len(death_claims), "type": "count"},
        {"label": "Total Amount Requested", "value": total_amount, "type": "currency"},
    ]
    report["columns"] = [
        _col("member_name", "Member", "left"),
        _col("college", "College", "left"),
        _col("claim_type", "Type", "left"),
        _col("current_stage", "Current Stage", "left"),
        _col("amount_requested", "Amount", "right"),
        _col("date_filed", "Date Filed", "left"),
        _col("status", "Status", "left"),
    ]
    report["rows"] = [
        {
            "member_name": c["member_name"],
            "college": c["college"],
            "claim_type": c["claim_type"],
            "current_stage": c["current_stage"],
            "amount_requested": c["amount_requested"],
            "date_filed": c["date_filed"],
            "status": c["status"],
        }
        for c in all_claims
    ]

    legacy = {
        "filters": {"claim_type": claim_type, "college": college},
        "summary": {
            "total_claims": len(all_claims),
            "medical_claims": len(medical_claims),
            "death_claims": len(death_claims),
            "total_amount_requested": total_amount,
        },
        "claims": all_claims,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 4. Membership Summary
# ---------------------------------------------------------------------------

def build_membership_summary(request):
    college = _get(request, "college")
    members = _active_members()
    if college:
        members = members.filter(department__iexact=college)

    total_members = members.count()
    current_year = timezone.now().year

    active_members = members.filter(
        Q(membership_status__iexact="active") | Q(employment_status__iexact="active")
    ).count()
    inactive_members = total_members - active_members
    new_members_this_year = members.filter(date_joined__year=current_year).count()

    college_breakdown = []
    for dept in members.values_list("department", flat=True).distinct().order_by("department"):
        if not dept:
            continue
        count = members.filter(department=dept).count()
        college_breakdown.append({
            "college": dept,
            "count": count,
            "percentage": round((count / total_members * 100), 1) if total_members else 0,
        })

    type_breakdown = [
        {"type": item["membership_status"] or "Unknown", "count": item["count"]}
        for item in members.values("membership_status").annotate(count=Count("member_id_PK")).order_by("-count")
    ]

    yearly_trend = []
    for year in [current_year - 2, current_year - 1, current_year]:
        yearly_trend.append({"year": year, "count": members.filter(date_joined__year=year).count()})

    filters = [{"label": "College", "value": college or "All"}]

    report = _report_base(
        "membership_summary", "Membership Summary",
        "Overall membership statistics, distribution by college and yearly joining trend.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Total Registered", "value": total_members, "type": "count"},
        {"label": "New This Year", "value": new_members_this_year, "type": "count"},
        {"label": "Active", "value": active_members, "type": "count"},
        {"label": "Inactive", "value": inactive_members, "type": "count"},
    ]
    report["columns"] = [
        _col("college", "College", "left"),
        _col("count", "Members", "right"),
        _col("percentage", "% of Total", "right"),
    ]
    report["rows"] = college_breakdown
    report["sections"] = [
        {
            "title": "By Membership Status",
            "columns": [_col("type", "Status", "left"), _col("count", "Count", "right")],
            "rows": type_breakdown,
        },
        {
            "title": "Membership Trend (Last 3 Years)",
            "columns": [_col("year", "Year", "left"), _col("count", "Members", "right")],
            "rows": yearly_trend,
        },
    ]

    legacy = {
        "filters": {"college": college},
        "summary": {
            "total_registered": total_members,
            "new_members_this_year": new_members_this_year,
            "active_members": active_members,
            "inactive_members": inactive_members,
        },
        "by_college": college_breakdown,
        "by_membership_type": type_breakdown,
        "yearly_trend": yearly_trend,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 5. Membership Status
# ---------------------------------------------------------------------------

def build_membership_status(request):
    college = _get(request, "college")
    members = _active_members()
    if college:
        members = members.filter(department__iexact=college)

    status_data: "OrderedDict[str, dict]" = OrderedDict()
    for member in members.order_by("membership_status", "full_name"):
        status = member.membership_status or "Unknown"
        entry = status_data.setdefault(status, {"status": status, "count": 0, "members": []})
        entry["count"] += 1
        entry["members"].append({
            "member_id": member.member_id_PK,
            "full_name": member.full_name,
            "position": member.position or "N/A",
            "college": member.department or "N/A",
            "department": member.department or "N/A",
            "date_joined": member.date_joined.strftime("%Y-%m-%d") if member.date_joined else "N/A",
        })

    total_members = sum(e["count"] for e in status_data.values())

    filters = [{"label": "College", "value": college or "All"}]

    report = _report_base(
        "membership_status", "Membership Status",
        "Members grouped by membership status with a breakdown per status group.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Total Members", "value": total_members, "type": "count"},
        {"label": "Status Categories", "value": len(status_data), "type": "count"},
    ]
    report["columns"] = [
        _col("status", "Status", "left"),
        _col("count", "Count", "right"),
        _col("percentage", "% of Total", "right"),
    ]
    report["rows"] = [
        {
            "status": e["status"],
            "count": e["count"],
            "percentage": round((e["count"] / total_members * 100), 1) if total_members else 0,
        }
        for e in status_data.values()
    ]
    report["sections"] = [
        {
            "title": f"{e['status']} ({e['count']} members)",
            "columns": [
                _col("full_name", "Name", "left"),
                _col("position", "Position", "left"),
                _col("department", "Department", "left"),
                _col("date_joined", "Date Joined", "left"),
            ],
            "rows": e["members"],
        }
        for e in status_data.values()
    ]

    legacy = {
        "filters": {"college": college},
        "report_data": list(status_data.values()),
        "summary": {"total_members": total_members, "status_categories": len(status_data)},
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 6. Monthly Dues Summary
# ---------------------------------------------------------------------------

def build_monthly_dues_summary(request):
    year = _get(request, "year") or str(timezone.now().year)
    college = _get(request, "college")

    members = _active_members()
    if college:
        members = members.filter(department__iexact=college)
    total_members = members.count()

    dues = MonthlyDues.objects.filter(month_covered__startswith=year)
    if college:
        dues = dues.filter(member_id_FK__department__iexact=college)

    monthly_dues_amount = get_monthly_dues_amount()

    monthly_breakdown = []
    for month in range(1, 13):
        month_key = f"{year}-{str(month).zfill(2)}"
        month_dues = dues.filter(month_covered=month_key)
        collected = month_dues.filter(payment_status__in=list(PAID_STATUSES)).aggregate(total=Sum("amount"))["total"] or 0
        expected = total_members * monthly_dues_amount
        monthly_breakdown.append({
            "month": month_key,
            "month_name": MONTH_NAMES[month - 1],
            "expected": _as_float(expected),
            "collected": _as_float(collected),
            "unpaid": max(_as_float(expected) - _as_float(collected), 0),
        })

    payment_method_rows = []
    for item in dues.values("payment_method").annotate(count=Sum("amount")).order_by("-count"):
        if not item["payment_method"]:
            continue
        method_dues = dues.filter(payment_method=item["payment_method"])
        payment_method_rows.append({
            "method": item["payment_method"],
            "members": method_dues.count(),
            "amount": _as_float(method_dues.aggregate(total=Sum("amount"))["total"]),
        })

    total_expected = total_members * monthly_dues_amount * 12
    total_collected = dues.filter(payment_status__in=list(PAID_STATUSES)).aggregate(total=Sum("amount"))["total"] or 0

    filters = [
        {"label": "Year", "value": year},
        {"label": "College", "value": college or "All"},
    ]

    report = _report_base(
        "monthly_dues_summary", "Monthly Dues Summary",
        "Monthly dues collection across the year with a payment method breakdown.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Expected per Member / Month", "value": _as_float(monthly_dues_amount), "type": "currency"},
        {"label": "Total Expected", "value": _as_float(total_expected), "type": "currency"},
        {"label": "Total Collected", "value": _as_float(total_collected), "type": "currency"},
        {"label": "Collection Rate", "value": round((_as_float(total_collected) / _as_float(total_expected) * 100) if _as_float(total_expected) else 0, 2), "type": "percent"},
    ]
    report["columns"] = [
        _col("month_name", "Month", "left"),
        _col("expected", "Expected", "right"),
        _col("collected", "Collected", "right"),
        _col("unpaid", "Unpaid", "right"),
    ]
    report["rows"] = monthly_breakdown
    report["sections"] = [
        {
            "title": "Payment Method",
            "columns": [
                _col("method", "Payment Method", "left"),
                _col("members", "Members", "right"),
                _col("amount", "Amount", "right"),
            ],
            "rows": payment_method_rows,
        },
    ]

    legacy = {
        "filters": {"year": year, "college": college},
        "summary": {
            "expected_per_month": _as_float(monthly_dues_amount),
            "total_expected": _as_float(total_expected),
            "total_collected": _as_float(total_collected),
            "total_unpaid": _as_float(total_expected) - _as_float(total_collected),
        },
        "monthly_breakdown": monthly_breakdown,
        "payment_methods": payment_method_rows,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 7. Contributions Summary
# ---------------------------------------------------------------------------

def build_contributions_summary(request):
    year = _get(request, "year") or str(timezone.now().year)

    posts = AidTrackingPost.objects.filter(Q(total_expected__gt=0) | Q(total_collected__gt=0))
    posts = posts.filter(target_month__startswith=year)

    contributions_table = []
    medical_expected = medical_collected = 0.0
    death_expected = death_collected = 0.0
    total_expected = total_collected = 0.0

    for post in posts:
        expected = _as_float(post.total_expected)
        collected = _as_float(post.total_collected)
        if not expected:
            agg = Contribution.objects.filter(aid_tracking_post_id_FK=post).aggregate(
                exp=Sum("expected_amount"), col=Sum("paid_amount")
            )
            expected = _as_float(agg["exp"])
            collected = _as_float(agg["col"])

        label = AID_TYPE_LABELS.get(post.aid_type, post.aid_type)
        total_expected += expected
        total_collected += collected
        if post.aid_type == "medical_aid":
            medical_expected += expected
            medical_collected += collected
        else:
            death_expected += expected
            death_collected += collected

        contributions_table.append({
            "case_id": post.post_id_PK,
            "type": label,
            "expected": expected,
            "collected": collected,
            "remaining": max(expected - collected, 0),
            "status": "Complete" if collected >= expected else "Collecting",
        })

    filters = [{"label": "Year", "value": year}]

    report = _report_base(
        "contributions_summary", "Contributions Summary",
        "Expected and collected aid contributions per tracking post for the selected year.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Total Expected", "value": total_expected, "type": "currency"},
        {"label": "Total Collected", "value": total_collected, "type": "currency"},
        {"label": "Medical Aid Expected", "value": medical_expected, "type": "currency"},
        {"label": "Medical Aid Collected", "value": medical_collected, "type": "currency"},
        {"label": "Death Aid Expected", "value": death_expected, "type": "currency"},
        {"label": "Death Aid Collected", "value": death_collected, "type": "currency"},
    ]
    report["columns"] = [
        _col("case_id", "Aid Case", "left"),
        _col("type", "Type", "left"),
        _col("expected", "Expected", "right"),
        _col("collected", "Collected", "right"),
        _col("remaining", "Remaining", "right"),
        _col("status", "Status", "left"),
    ]
    report["rows"] = contributions_table

    legacy = {
        "filters": {"year": year},
        "summary": {
            "total_contributions": total_expected,
            "total_collected": total_collected,
            "medical_expected": medical_expected,
            "medical_collected": medical_collected,
            "death_expected": death_expected,
            "death_collected": death_collected,
        },
        "contributions_table": contributions_table,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 8. Fund Summary
# ---------------------------------------------------------------------------

def build_fund_summary(request):
    year = _get(request, "year") or str(timezone.now().year)
    int_year = int(year)

    opening_balance = (
        FundTransaction.objects.filter(recorded_at__year__lt=int_year)
        .aggregate(total=Sum("amount", filter=Q(direction="inflow")) - Sum("amount", filter=Q(direction="outflow")))
        ["total"] or 0
    )

    year_txs = FundTransaction.objects.filter(recorded_at__year=int_year).order_by("recorded_at")
    total_inflow = year_txs.filter(direction="inflow").aggregate(total=Sum("amount"))["total"] or 0
    total_outflow = year_txs.filter(direction="outflow").aggregate(total=Sum("amount"))["total"] or 0
    current_balance = _as_float(opening_balance) + _as_float(total_inflow) - _as_float(total_outflow)

    running = _as_float(opening_balance)
    fund_activity = []
    for tx in year_txs:
        inflow = _as_float(tx.amount) if tx.direction == "inflow" else 0
        outflow = _as_float(tx.amount) if tx.direction == "outflow" else 0
        running += inflow - outflow
        fund_activity.append({
            "date": tx.recorded_at.strftime("%Y-%m-%d %H:%M") if tx.recorded_at else "N/A",
            "description": tx.description or tx.source_type or "Transaction",
            "inflow": inflow,
            "outflow": outflow,
            "balance": running,
        })

    filters = [{"label": "Year", "value": year}]

    report = _report_base(
        "fund_summary", "Fund Summary",
        "Fund position showing opening balance, inflows, outflows and the running balance.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Opening Balance", "value": _as_float(opening_balance), "type": "currency"},
        {"label": "Total Inflow", "value": _as_float(total_inflow), "type": "currency"},
        {"label": "Total Outflow", "value": _as_float(total_outflow), "type": "currency"},
        {"label": "Current Balance", "value": current_balance, "type": "currency"},
    ]
    report["columns"] = [
        _col("date", "Date", "left"),
        _col("description", "Description", "left"),
        _col("inflow", "Inflow", "right"),
        _col("outflow", "Outflow", "right"),
        _col("balance", "Balance", "right"),
    ]
    report["rows"] = fund_activity

    legacy = {
        "filters": {"year": year},
        "summary": {
            "opening_balance": _as_float(opening_balance),
            "total_inflow": _as_float(total_inflow),
            "total_outflow": _as_float(total_outflow),
            "current_balance": current_balance,
        },
        "fund_activity": fund_activity,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 9. Medical Aid
# ---------------------------------------------------------------------------

def _claim_status_bucket(status, president_decision, released):
    if released:
        return "Released"
    if status in ("Rejected", "Returned for Revision"):
        return "Rejected"
    if president_decision == "Approved" or status in Status.ALL_APPROVED:
        return "Approved"
    if status in PENDING_STATUSES or status in ("Under Review", "Auditor Verified"):
        return "Pending"
    return status or "Unknown"


def build_medical_aid(request):
    year = _get(request, "year") or str(timezone.now().year)

    cases = MedicalAid.objects.filter(request_date__year=int(year))
    release_dates = _release_info_by_post()

    medical_table = []
    expected_per_member = get_contribution_amount_for_aid("Medical Aid")

    for case in cases:
        contributions = Contribution.objects.filter(
            aid_tracking_post_id_FK__source_type="medical_aid",
            aid_tracking_post_id_FK__source_id=str(case.medical_aid_id_PK),
        ).aggregate(total=Sum("paid_amount"))["total"] or 0
        released = case.released_by_user_id_FK_id is not None or case.status in ("Released", "Completed")

        medical_table.append({
            "case_id": case.medical_aid_id_PK,
            "member": case.member_id_FK.full_name if case.member_id_FK else "N/A",
            "bill_amount": _as_float(case.hospital_bill_amount),
            "expected_contribution": _as_float(expected_per_member),
            "contributions": _as_float(contributions),
            "released": _as_float(case.validated_aid_amount) if released else 0,
            "status": _claim_status_bucket(case.status, case.president_decision, released),
            "date": case.request_date.strftime("%Y-%m-%d") if case.request_date else "N/A",
        })

    def count(bucket):
        return sum(1 for c in medical_table if c["status"] == bucket)

    filters = [{"label": "Year", "value": year}]

    report = _report_base(
        "medical_aid", "Medical Aid",
        "Medical aid cases for the selected year with status and contribution figures.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Total Requests", "value": len(medical_table), "type": "count"},
        {"label": "Pending", "value": count("Pending"), "type": "count"},
        {"label": "Approved", "value": count("Approved"), "type": "count"},
        {"label": "Released", "value": count("Released"), "type": "count"},
        {"label": "Rejected", "value": count("Rejected"), "type": "count"},
    ]
    report["columns"] = [
        _col("member", "Member", "left"),
        _col("bill_amount", "Bill Amount", "right"),
        _col("expected_contribution", "Expected / Member", "right"),
        _col("contributions", "Contributions", "right"),
        _col("released", "Released", "right"),
        _col("status", "Status", "left"),
        _col("date", "Date", "left"),
    ]
    report["rows"] = medical_table

    legacy = {
        "filters": {"year": year},
        "summary": {
            "total_requests": len(medical_table),
            "pending": count("Pending"),
            "approved": count("Approved"),
            "released": count("Released"),
            "rejected": count("Rejected"),
        },
        "medical_table": medical_table,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 10. Death Aid
# ---------------------------------------------------------------------------

def build_death_aid(request):
    year = _get(request, "year") or str(timezone.now().year)

    cases = DeathAid.objects.filter(claim_date__year=int(year))

    death_table = []
    for case in cases:
        contributions = Contribution.objects.filter(
            aid_tracking_post_id_FK__source_type="death_aid",
            aid_tracking_post_id_FK__source_id=str(case.death_aid_id_PK),
        ).aggregate(total=Sum("paid_amount"))["total"] or 0
        released = case.released_by_user_id_FK_id is not None or case.status in ("Released", "Completed")
        category = DEATH_CATEGORY_LABELS.get(case.relationship_group, "Other")
        contribution = _as_float(case.benefit_amount) or {
            "member": 500, "spouse": 300, "parent_child": 250, "sibling": 100
        }.get(case.relationship_group, 0)

        death_table.append({
            "case_id": case.death_aid_id_PK,
            "requester": case.member_id_FK.full_name if case.member_id_FK else "N/A",
            "beneficiary": case.deceased_name or "N/A",
            "category": category,
            "contribution": contribution,
            "contributions": _as_float(contributions),
            "status": _claim_status_bucket(case.status, case.president_decision, released),
            "date": case.claim_date.strftime("%Y-%m-%d") if case.claim_date else "N/A",
        })

    def count(bucket):
        return sum(1 for c in death_table if c["status"] == bucket)

    filters = [{"label": "Year", "value": year}]

    report = _report_base(
        "death_aid", "Death Aid",
        "Death aid cases for the selected year with beneficiary category and benefit amount.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Total Requests", "value": len(death_table), "type": "count"},
        {"label": "Pending", "value": count("Pending"), "type": "count"},
        {"label": "Approved", "value": count("Approved"), "type": "count"},
        {"label": "Released", "value": count("Released"), "type": "count"},
        {"label": "Rejected", "value": count("Rejected"), "type": "count"},
    ]
    report["columns"] = [
        _col("requester", "Requester", "left"),
        _col("beneficiary", "Beneficiary", "left"),
        _col("category", "Category", "left"),
        _col("contribution", "Contribution", "right"),
        _col("contributions", "Contributions", "right"),
        _col("status", "Status", "left"),
        _col("date", "Date", "left"),
    ]
    report["rows"] = death_table
    report["notes"] = [
        "Member: \u20b1500  |  Husband/Wife: \u20b1300  |  Parents and Children: \u20b1250  |  Brother/Sister (Full Blood): \u20b1100",
    ]

    legacy = {
        "filters": {"year": year},
        "summary": {
            "total_requests": len(death_table),
            "pending": count("Pending"),
            "approved": count("Approved"),
            "released": count("Released"),
            "rejected": count("Rejected"),
        },
        "death_table": death_table,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 11. Approved Claims
# ---------------------------------------------------------------------------

def build_approved_claims(request):
    year = _get(request, "year") or str(timezone.now().year)

    approved_table = []

    for table_name, qs in (
        ("medical_aid", MedicalAid.objects.filter(request_date__year=int(year))),
        ("death_aid", DeathAid.objects.filter(claim_date__year=int(year))),
    ):
        for claim in qs:
            approved = (
                claim.president_decision == "Approved"
                or claim.status in Status.ALL_APPROVED
                or claim.status in ("Completed", "Released")
            )
            if not approved:
                continue
            treasurer = "Approved" if claim.treasurer_validated_by_user_id_FK_id else "Not Approved"
            auditor = "Approved" if claim.auditor_verified_by_user_id_FK_id else "Not Approved"
            president = "Approved" if claim.president_decision == "Approved" else "Not Approved"
            type_label = "Medical Aid" if table_name == "medical_aid" else "Death Aid"
            claim_id = getattr(claim, "medical_aid_id_PK", None) or getattr(claim, "death_aid_id_PK")
            member = claim.member_id_FK.full_name if claim.member_id_FK else "N/A"

            approved_table.append({
                "case_id": claim_id,
                "type": type_label,
                "requester": member,
                "treasurer": treasurer,
                "auditor": auditor,
                "president": president,
                "approval_date": _latest_approval_datetime(table_name, claim_id),
            })

    approved_table.sort(key=lambda c: c["case_id"])

    filters = [{"label": "Year", "value": year}]

    report = _report_base(
        "approved_claims", "Approved Claims",
        "Claims that have completed the Treasurer, Auditor and President approval chain.",
        request, filters,
    )
    report["summary"] = [{"label": "Total Approved", "value": len(approved_table), "type": "count"}]
    report["columns"] = [
        _col("type", "Type", "left"),
        _col("requester", "Requester", "left"),
        _col("treasurer", "Treasurer", "center"),
        _col("auditor", "Auditor", "center"),
        _col("president", "President", "center"),
        _col("approval_date", "Approval Date", "left"),
    ]
    report["rows"] = approved_table

    legacy = {
        "filters": {"year": year},
        "summary": {"total_approved": len(approved_table)},
        "approved_table": approved_table,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# 12. Released Claims
# ---------------------------------------------------------------------------

def build_released_claims(request):
    year = _get(request, "year") or str(timezone.now().year)
    release_dates = _release_info_by_post()

    released_table = []
    total_medical_released = Decimal_0 = 0.0
    total_death_released = 0.0

    for table_name, qs in (
        ("medical_aid", MedicalAid.objects.filter(request_date__year=int(year))),
        ("death_aid", DeathAid.objects.filter(claim_date__year=int(year))),
    ):
        for claim in qs:
            release_date = "N/A"
            released = claim.released_by_user_id_FK_id is not None or claim.status in ("Released", "Completed")

            # Cross-check via tracking post release audit trail.
            post = AidTrackingPost.objects.filter(
                source_type=table_name, source_id=str(getattr(claim, "medical_aid_id_PK", None) or getattr(claim, "death_aid_id_PK"))
            ).first()
            if post and post.post_id_PK in release_dates:
                release_date = release_dates[post.post_id_PK]
                released = True

            if not released:
                continue

            amount = _as_float(claim.validated_aid_amount if table_name == "medical_aid" else claim.benefit_amount)
            type_label = "Medical Aid" if table_name == "medical_aid" else "Death Aid"
            if table_name == "medical_aid":
                total_medical_released += amount
            else:
                total_death_released += amount

            released_table.append({
                "case_id": getattr(claim, "medical_aid_id_PK", None) or getattr(claim, "death_aid_id_PK"),
                "type": type_label,
                "requester": claim.member_id_FK.full_name if claim.member_id_FK else "N/A",
                "approved_amount": amount,
                "released_amount": amount,
                "release_date": release_date,
                "status": "Released",
            })

    released_table.sort(key=lambda c: c["case_id"])

    filters = [{"label": "Year", "value": year}]

    report = _report_base(
        "released_claims", "Released Claims",
        "Claims that have been released to the requester with the released amounts.",
        request, filters,
    )
    report["summary"] = [
        {"label": "Medical Aid Released", "value": total_medical_released, "type": "currency"},
        {"label": "Death Aid Released", "value": total_death_released, "type": "currency"},
        {"label": "Total Released", "value": total_medical_released + total_death_released, "type": "currency"},
    ]
    report["columns"] = [
        _col("type", "Type", "left"),
        _col("requester", "Requester", "left"),
        _col("approved_amount", "Approved Amount", "right"),
        _col("released_amount", "Released Amount", "right"),
        _col("release_date", "Release Date", "left"),
        _col("status", "Status", "left"),
    ]
    report["rows"] = released_table

    legacy = {
        "filters": {"year": year},
        "summary": {
            "total_medical_released": total_medical_released,
            "total_death_released": total_death_released,
            "total_released": total_medical_released + total_death_released,
        },
        "released_table": released_table,
        "generated_at": _now_str(),
    }
    return report, legacy


# ---------------------------------------------------------------------------
# Dispatch
# ---------------------------------------------------------------------------

BUILDERS = {
    "members_by_college": build_members_by_college,
    "paid_unpaid_summary": build_paid_unpaid_summary,
    "pending_claims": build_pending_claims,
    "membership_summary": build_membership_summary,
    "membership_status": build_membership_status,
    "monthly_dues_summary": build_monthly_dues_summary,
    "contributions_summary": build_contributions_summary,
    "fund_summary": build_fund_summary,
    "medical_aid": build_medical_aid,
    "death_aid": build_death_aid,
    "approved_claims": build_approved_claims,
    "released_claims": build_released_claims,
}


def build_report(request, report_key):
    """Return (report, legacy) for a report key or (None, None) if unknown."""
    builder = BUILDERS.get(report_key)
    if builder is None:
        return None, None
    return builder(request)


def available_departments():
    """Distinct department strings used to populate college filter dropdowns."""
    return sorted(
        (d for d in Member.objects.exclude(department__isnull=True).values_list("department", flat=True).distinct() if d),
        key=str,
    )


def available_years():
    """Years that have data across member/dues/claims tables."""
    years = set()
    for qs, field in (
        (Member.objects.all(), "date_joined"),
        (MonthlyDues.objects.all(), "month_covered"),
        (MedicalAid.objects.all(), "request_date"),
        (DeathAid.objects.all(), "claim_date"),
        (FundTransaction.objects.all(), "recorded_at"),
    ):
        for row in qs.values_list(field, flat=True).distinct():
            if row:
                try:
                    years.add(int(str(row)[:4]))
                except (ValueError, TypeError):
                    continue
    if not years:
        years.add(timezone.now().year)
    return sorted(years)


# ---------------------------------------------------------------------------
# Export serializers (Excel / PDF)
# ---------------------------------------------------------------------------

EXPORT_TITLES = {
    "members_by_college": "Members by College",
    "paid_unpaid_summary": "Paid / Unpaid Summary",
    "pending_claims": "Pending Claims",
    "membership_summary": "Membership Summary",
    "membership_status": "Membership Status",
    "monthly_dues_summary": "Monthly Dues Summary",
    "contributions_summary": "Contributions Summary",
    "fund_summary": "Fund Summary",
    "medical_aid": "Medical Aid",
    "death_aid": "Death Aid",
    "approved_claims": "Approved Claims",
    "released_claims": "Released Claims",
}


def _safe_sheet_name(name):
    import re

    cleaned = re.sub(r"[\[\]:*?/\\]", "_", name or "Report")[:31]
    return cleaned or "Report"


def report_to_xlsx(report):
    """Build an openpyxl Workbook from a normalized report dict."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(report.get("report_name") or report.get("report_key"))

    title = report.get("report_name") or EXPORT_TITLES.get(report.get("report_key"), "Report")
    ws["A1"] = f"ISU - CAUFA: {title}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    meta = []
    meta.append(f"Generated By: {report.get('generated_by')}")
    meta.append(f"Generated At: {report.get('generated_at')}")
    for f in report.get("filters", []):
        meta.append(f"{f.get('label')}: {f.get('value')}")
    row = 2
    for m in meta:
        ws.cell(row=row, column=1, value=m).font = Font(size=9, italic=True)
        row += 1

    row += 1  # blank line
    start = row

    # Summary block
    summ = report.get("summary", [])
    if summ:
        for i, s in enumerate(summ):
            ws.cell(row=row, column=1 + (i % 2) * 3, value=s.get("label")).font = Font(bold=True, size=10)
            ws.cell(row=row, column=2 + (i % 2) * 3, value=s.get("value"))
            if (i + 1) % 2 == 0:
                row += 1
        if len(summ) % 2:
            row += 1
        row += 1

    def render_table(ws, row, columns, rows):
        header_fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
        header_font = Font(color="ffffff", bold=True, size=10)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, col in enumerate(columns, 1):
            cell = ws.cell(row=row, column=ci, value=col.get("label"))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        row += 1
        for r in rows:
            for ci, col in enumerate(columns, 1):
                cell = ws.cell(row=row, column=ci, value=r.get(col.get("key")))
                cell.border = border
                cell.alignment = Alignment(horizontal=col.get("align", "left"))
            row += 1
        return row

    row = render_table(ws, row, report.get("columns", []), report.get("rows", []))

    for section in report.get("sections", []):
        row += 1
        ws.cell(row=row, column=1, value=section.get("title", "")).font = Font(bold=True, size=11)
        row += 1
        row = render_table(ws, row, section.get("columns", []), section.get("rows", []))

    notes = report.get("notes", [])
    if notes:
        row += 1
        for n in notes:
            ws.cell(row=row, column=1, value=n).font = Font(size=9, italic=True)
            row += 1

    for col_idx in range(1, min(len(report.get("columns", [])) or 1, 6) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 22
    return wb


def report_to_pdf(report):
    """Build a PDF that mirrors the report preview layout using the same report schema."""
    import html

    def _fallback_reportlab():
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
        import io

        GREEN_DARK = colors.HexColor("#0f5e3d")
        GREEN = colors.HexColor("#2e7d32")
        GREEN_LIGHT_BG = colors.HexColor("#dfeee5")
        CARD_BG = colors.HexColor("#f0fdf4")
        CARD_BORDER = colors.HexColor("#bbf7d0")
        CARD_LABEL = colors.HexColor("#15803d")
        CARD_VALUE = colors.HexColor("#166534")
        PILL_BG = colors.HexColor("#f3f4f6")
        PILL_BORDER = colors.HexColor("#e5e7eb")
        PILL_TEXT = colors.HexColor("#4b5563")
        TEXT_DARK = colors.HexColor("#111827")
        CELL_TEXT = colors.HexColor("#1f2937")
        ALT_ROW = colors.HexColor("#f9fafb")
        GRID_BORDER = colors.HexColor("#d0d7de")
        SHEET_BORDER = colors.HexColor("#d9d9d9")

        report_name = report.get("report_name") or EXPORT_TITLES.get(report.get("report_key"), "Report")
        generated_by = report.get("generated_by") or "President"
        generated_at = report.get("generated_at") or ""

        def fmt_summary_value(item):
            value = item.get("value", "")
            if item.get("type") == "currency":
                try:
                    return f"\u20b1{float(value):,.2f}"
                except Exception:
                    return str(value)
            return str(value)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            rightMargin=12 * mm,
            leftMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )

        base = ParagraphStyle("Base", fontName="Helvetica", fontSize=9, textColor=CELL_TEXT)
        brand_style = ParagraphStyle("Brand", parent=base, fontName="Helvetica-Bold", fontSize=24, leading=24, textColor=TEXT_DARK)
        brand_sub_style = ParagraphStyle("BrandSub", parent=base, fontSize=7.5, leading=9, textColor=PILL_TEXT)
        meta_style = ParagraphStyle("Meta", parent=base, fontSize=8.5, leading=12, textColor=PILL_TEXT, alignment=2)
        title_style = ParagraphStyle(
            "TitleX", parent=base, fontName="Helvetica-Bold", fontSize=18, leading=22,
            alignment=1, textColor=TEXT_DARK, spaceBefore=10, spaceAfter=6,
        )
        pill_style = ParagraphStyle(
            "Pill", parent=base, fontSize=8.5, leading=11, textColor=PILL_TEXT,
            backColor=PILL_BG, borderColor=PILL_BORDER, borderWidth=0.6,
            borderPadding=3, borderRadius=4,
        )
        card_label_style = ParagraphStyle("CardLabel", parent=base, fontName="Helvetica-Bold", fontSize=7.5, leading=9, textColor=CARD_LABEL)
        card_value_style = ParagraphStyle("CardValue", parent=base, fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=CARD_VALUE)
        sec_style = ParagraphStyle(
            "Section", parent=base, fontName="Helvetica-Bold", fontSize=10, leading=12,
            textColor=colors.HexColor("#123b2a"),
        )
        note_style = ParagraphStyle("Note", parent=base, fontSize=8, leading=10, textColor=PILL_TEXT)
        cell_style = ParagraphStyle("Cell", parent=base, fontSize=8, leading=10, textColor=CELL_TEXT)
        header_cell_style = ParagraphStyle("HeaderCell", parent=base, fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.white)

        story = []

        # --- Brand header row: logo + name + subtitle (left), meta (right) ---
        logo_flow = None
        logo_path = Path(settings.BASE_DIR) / "static" / "img" / "isu_caufa_official.png"
        if logo_path.exists():
            try:
                logo_flow = Image(str(logo_path), width=14 * mm, height=14 * mm)
            except Exception:
                logo_flow = None

        brand_cell = []
        if logo_flow is not None:
            brand_cell.append(logo_flow)
        brand_cell.append(
            Paragraph(
                f"<b>ISU CAUFA</b><br/><font size=7.5 color='#4b5563'>ISABELA STATE UNIVERSITY \u2013 CAUAYAN CAMPUS FACULTY ASSOCIATION</font>",
                brand_style,
            )
        )
        header_left = Table([[brand_cell]], colWidths=[None])
        header_left.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        header_right = Table(
            [[
                Paragraph(f"<b>Date Generated:</b> {html.escape(str(generated_at))}", meta_style),
                Paragraph(f"<b>Generated By:</b> {html.escape(str(generated_by))}", meta_style),
            ]],
            colWidths=[None],
        )
        header_right.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ]))

        header = Table([[header_left, header_right]], colWidths=[int(doc.width * 0.62), int(doc.width * 0.38)])
        header.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 0), (-1, 0), 2, GREEN_DARK),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(header)

        # --- Title ---
        story.append(Paragraph(html.escape(str(report_name)).upper(), title_style))

        # --- Summary cards (2-column grid) ---
        summ = report.get("summary", [])
        if summ:
            cards = []
            for s in summ:
                card = Table(
                    [[Paragraph(html.escape(str(s.get("label", ""))).upper(), card_label_style)],
                     [Paragraph(html.escape(fmt_summary_value(s)), card_value_style)]],
                    colWidths=[int(doc.width / 2) - 6 * mm],
                )
                card.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), CARD_BG),
                    ("BOX", (0, 0), (-1, -1), 1, CARD_BORDER),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]))
                cards.append(card)
            card_grid_rows = []
            for i in range(0, len(cards), 2):
                row = cards[i:i + 2]
                while len(row) < 2:
                    row.append("")
                card_grid_rows.append(row)
            card_grid = Table(card_grid_rows, colWidths=[int(doc.width / 2) - 3 * mm, int(doc.width / 2) - 3 * mm])
            card_grid.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(card_grid)
            story.append(Spacer(1, 8))

        # --- Data table ---
        def build_table(columns, rows):
            if not columns:
                return None
            header_cells = [
                Paragraph(html.escape(str(c.get("label", ""))), header_cell_style) for c in columns
            ]
            data = [header_cells]
            if not rows:
                data.append([Paragraph("No records found", cell_style) for _ in columns])
            else:
                for row in rows:
                    cells = []
                    for c in columns:
                        key = c.get("key")
                        value = row.get(key, "") if isinstance(row, dict) else ""
                        cells.append(
                            Paragraph(
                                html.escape(str(value if value is not None else "")),
                                cell_style,
                            )
                        )
                    data.append(cells)
            tbl = Table(data, repeatRows=1)
            style = [
                ("BACKGROUND", (0, 0), (-1, 0), GREEN_DARK),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, GRID_BORDER),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ALT_ROW]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
            for idx, c in enumerate(columns):
                align = c.get("align", "left")
                if align == "right":
                    style.append(("ALIGN", (idx, 0), (idx, -1), "RIGHT"))
                elif align == "center":
                    style.append(("ALIGN", (idx, 0), (idx, -1), "CENTER"))
                else:
                    style.append(("ALIGN", (idx, 0), (idx, -1), "LEFT"))
            tbl.setStyle(TableStyle(style))
            return tbl

        main_table = build_table(report.get("columns", []), report.get("rows", []))
        if main_table is not None:
            story.append(main_table)

        # --- Sections ---
        for section in report.get("sections", []):
            story.append(Spacer(1, 6))
            sec_title = Table(
                [[Paragraph(f"<b>{html.escape(str(section.get('title', ''))).upper()}</b>", sec_style)]],
                colWidths=[doc.width],
            )
            sec_title.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), GREEN_LIGHT_BG),
                ("LINEBEFORE", (0, 0), (0, -1), 3, GREEN),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(sec_title)
            story.append(Spacer(1, 4))
            sec_table = build_table(section.get("columns", []), section.get("rows", []))
            if sec_table is not None:
                story.append(sec_table)

        # --- Notes ---
        notes = report.get("notes", [])
        if notes:
            story.append(Spacer(1, 6))
            for n in notes:
                story.append(Paragraph(html.escape(str(n)), note_style))

        doc.build(story)
        return buf.getvalue()

    try:
        from weasyprint import HTML
    except Exception:
        return _fallback_reportlab()

    def fmt_summary_value(item):
        value = item.get("value", "")
        if item.get("type") == "currency":
            try:
                return f"₱{float(value):,.2f}"
            except Exception:
                return str(value)
        return str(value)

    report_name = report.get("report_name") or EXPORT_TITLES.get(report.get("report_key"), "Report")
    generated_by = report.get("generated_by") or "President"
    generated_at = report.get("generated_at") or ""

    def render_filters(filters):
        if not filters:
            return ""
        parts = []
        for f in filters:
            label = html.escape(str(f.get("label", "")))
            value = html.escape(str(f.get("value", "")))
            parts.append(f'<span><b>{label}:</b> {value}</span>')
        return f'<div class="rpt-modal-filters">{"".join(parts)}</div>'

    def render_summary(summary):
        if not summary:
            return ""
        items = []
        for item in summary:
            label = html.escape(str(item.get("label", "")))
            value = html.escape(fmt_summary_value(item))
            items.append(
                f'<div class="rpt-modal-summary-item"><span>{label}</span><strong>{value}</strong></div>'
            )
        return f'<div class="rpt-modal-summary">{"".join(items)}</div>'

    def render_table(columns, rows):
        if not columns:
            return ""
        head = []
        for c in columns:
            label = html.escape(str(c.get("label", "")))
            align = c.get("align", "left")
            head.append(f'<th class="rpt-align-{align}">{label}</th>')

        body_rows = []
        if not rows:
            body_rows.append(f'<tr><td colspan="{len(columns)}" class="rpt-empty">No records found</td></tr>')
        else:
            for row in rows:
                cells = []
                for c in columns:
                    key = c.get("key")
                    value = row.get(key, "") if isinstance(row, dict) else ""
                    align = c.get("align", "left")
                    text = html.escape(str(value if value is not None else ""))
                    cells.append(f'<td class="rpt-align-{align}">{text}</td>')
                body_rows.append(f'<tr>{"".join(cells)}</tr>')

        return (
            '<div class="rpt-modal-table-wrap">'
            '<table class="rpt-modal-table">'
            f'<thead><tr>{"".join(head)}</tr></thead>'
            f'<tbody>{"".join(body_rows)}</tbody>'
            '</table></div>'
        )

    def render_sections(sections):
        if not sections:
            return ""
        output = []
        for sec in sections:
            title = html.escape(str(sec.get("title", "")))
            output.append(f'<h4 class="rpt-modal-section-title">{title}</h4>')
            output.append(render_table(sec.get("columns", []), sec.get("rows", [])))
        return "".join(output)

    logo_path = Path(settings.BASE_DIR) / "static" / "img" / "isu_caufa_official.png"
    logo_src = ""
    if logo_path.exists():
        try:
            logo_bytes = logo_path.read_bytes()
            logo_src = "data:image/png;base64," + base64.b64encode(logo_bytes).decode("ascii")
        except Exception:
            logo_src = "/static/img/isu_caufa_official.png"
    else:
        logo_src = "/static/img/isu_caufa_official.png"

    notes = report.get("notes") or []
    notes_html = "".join(f'<p>{html.escape(str(n))}</p>' for n in notes)
    if notes_html:
        notes_html = f'<div class="rpt-modal-notes">{notes_html}</div>'

    filters_html = render_filters(report.get("filters", []))
    summary_html = render_summary(report.get("summary", []))
    table_html = render_table(report.get("columns", []), report.get("rows", []))
    sections_html = render_sections(report.get("sections", []))

    html_doc = f'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <title>{html.escape(report_name)}</title>
  <style>
    @page {{ size: A4 landscape; margin: 10mm; }}
    html, body {{ margin: 0; padding: 0; background: #ffffff; font-family: Arial, Helvetica, sans-serif; color: #111827; }}
    body {{ padding: 0; }}
    .rpt-report-shell {{ width: 100%; box-sizing: border-box; }}
    .rpt-report-sheet {{ width: 100%; background: #fff; border: 1px solid #d9d9d9; box-sizing: border-box; padding: 16px 18px 12px; }}
    .rpt-sheet-header {{ display: flex; justify-content: space-between; align-items: center; border-bottom: 2px solid #0f5e3d; padding-bottom: 10px; margin-bottom: 8px; }}
    .rpt-sheet-brand {{ display: flex; align-items: center; gap: 12px; }}
    .rpt-logo-badge {{ width: 48px; height: 48px; border-radius: 50%; background: linear-gradient(135deg, #0e6b3b, #28a745); color: white; display: flex; align-items: center; justify-content: center; font-weight: 800; font-size: 18px; box-shadow: inset 0 0 0 3px rgba(255,255,255,0.6); }}
    .rpt-brand-name {{ font-size: 29px; line-height: 1; font-weight: 800; letter-spacing: 0.04em; color: #111827; }}
    .rpt-brand-sub {{ font-size: 10px; letter-spacing: 0.08em; color: #4b5563; margin-top: 4px; }}
    .rpt-sheet-meta {{ font-size: 12px; color: #374151; text-align: right; line-height: 1.4; }}
    .rpt-sheet-title {{ font-size: 26px; font-weight: 800; letter-spacing: 0.06em; text-transform: uppercase; text-align: center; margin: 16px 0 12px; color: #111827; }}
    .rpt-modal-filters {{ display: flex; flex-wrap: wrap; gap: 8px 12px; border-top: 1px solid #d0d7de; border-bottom: 1px solid #d0d7de; padding: 10px 0; margin: 0 0 16px; font-size: 12px; color: #4b5563; }}
    .rpt-modal-filters span {{ background: #f3f4f6; border: 1px solid #e5e7eb; padding: 4px 8px; border-radius: 4px; }}
    .rpt-modal-summary {{ display: grid; grid-template-columns: repeat(2, minmax(180px, 1fr)); gap: 12px; margin: 0 0 18px; }}
    .rpt-modal-summary-item {{ background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 8px; padding: 10px 12px; }}
    .rpt-modal-summary-item span {{ display: block; font-size: 11px; color: #15803d; font-weight: 700; text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 4px; }}
    .rpt-modal-summary-item strong {{ font-size: 18px; color: #166534; }}
    .rpt-modal-table-wrap {{ overflow: hidden; border: 1px solid #d0d7de; margin-bottom: 16px; }}
    @media print {{
      .rpt-sheet-brand img {{ width: 72px !important; height: 72px !important; max-width: 72px !important; max-height: 72px !important; }}
    }}
    .rpt-modal-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
    .rpt-modal-table th {{ background: #0f5e3d; color: white; padding: 8px 10px; text-align: left; font-weight: 700; border-right: 1px solid rgba(255,255,255,0.15); }}
    .rpt-modal-table td {{ padding: 8px 10px; border-top: 1px solid #e5e7eb; color: #1f2937; border-right: 1px solid #f0f0f0; }}
    .rpt-modal-table tbody tr:nth-child(even) {{ background: #f9fafb; }}
    .rpt-align-left {{ text-align: left; }}
    .rpt-align-right {{ text-align: right; }}
    .rpt-align-center {{ text-align: center; }}
    .rpt-empty {{ text-align: center; color: #6b7280; padding: 28px; }}
    .rpt-modal-section-title {{ margin: 18px 0 10px; font-size: 14px; line-height: 1.3; color: #123b2a; background: #dfeee5; border-left: 4px solid #2e7d32; padding: 8px 10px; text-transform: uppercase; font-weight: 800; }}
    .rpt-modal-notes {{ margin-top: 12px; font-size: 11px; color: #4b5563; }}
    .rpt-modal-notes p {{ margin: 4px 0; }}
  </style>
</head>
<body>
  <div class="rpt-report-shell">
    <div class="rpt-report-sheet">
      <div class="rpt-sheet-header">
        <div class="rpt-sheet-brand">
          <img src="{logo_src}" alt="ISU CAUFA Logo" style="width: 52px; height: 52px; object-fit: contain; border-radius: 50%; background: #f4f7f5; border: 1px solid #d9e5dd;" />
          <div>
            <div class="rpt-brand-name">ISU CAUFA</div>
            <div class="rpt-brand-sub">ISABELA STATE UNIVERSITY – CAUAYAN CAMPUS FACULTY ASSOCIATION</div>
          </div>
        </div>
        <div class="rpt-sheet-meta">
          <div>Date Generated: {html.escape(generated_at)}</div>
          <div>Generated By: {html.escape(generated_by)}</div>
        </div>
      </div>
      <div class="rpt-sheet-title">{html.escape(report_name.upper())}</div>
      {summary_html}
      {table_html}
      {sections_html}
      {notes_html}
    </div>
  </div>
</body>
</html>'''

    try:
        pdf = HTML(string=html_doc, base_url="").write_pdf()
    except Exception:
        return _fallback_reportlab()
    return pdf

    