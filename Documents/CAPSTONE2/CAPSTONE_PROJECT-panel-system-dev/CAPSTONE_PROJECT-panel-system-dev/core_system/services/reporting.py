from __future__ import annotations

import os
import pytz
from datetime import datetime
from decimal import Decimal
from typing import Optional

from django.conf import settings
from django.core.files.storage import default_storage
from django.utils import timezone

from core_system.models import (
    Member,
    Department,
    MonthlyDues,
    MembershipFee,
    AidTrackingPost,
    Contribution,
    FundTransaction,
    OrganizationFundReport,
)
from core_system.services.compliance import (
    dues_compliance_summary,
    member_dues_status,
    member_contribution_status,
    dues_overdue_bucket,
)


def _style_header(ws, row, cols):
    from openpyxl.styles import Alignment, Font, PatternFill, Side, Border

    header_fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
    header_font = Font(color="ffffff", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _auto_width(ws, cols):
    for col in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[chr(64 + col)].width = max_len + 4


def _fmt(num):
    if num is None:
        return 0.0
    return round(float(num), 2)


def generate_overall_report(
    year: int | None = None,
    month: int | None = None,
    wb=None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    today = timezone.localdate()
    year = year or today.year
    month = month or today.month
    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Overall Summary"
    else:
        ws = wb.create_sheet("Overall Summary")
    ws.cell(row=1, column=1, value=f"Compliance Report - {year}-{month:02d}").font = Font(
        bold=True, size=14
    )
    ws.merge_cells("A1:E1")

    headers = ["Metric", "Value"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    _style_header(ws, 3, 2)

    dept_summary = dues_compliance_summary(year, month)
    total_members = sum(d["total_members"] for d in dept_summary)
    total_paid = sum(d["paid_count"] for d in dept_summary)
    total_unpaid = sum(d["unpaid_count"] for d in dept_summary)
    overall_pct = round(total_paid / total_members * 100, 1) if total_members else 0.0

    data_rows = [
        ("Period", f"{year}-{month:02d}"),
        ("Total Active Members", total_members),
        ("Total Paid", total_paid),
        ("Total Unpaid", total_unpaid),
        ("Compliance %", f"{overall_pct}%"),
    ]
    for ri, (metric, val) in enumerate(data_rows, 4):
        ws.cell(row=ri, column=1, value=metric)
        ws.cell(row=ri, column=2, value=val)
    _auto_width(ws, 2)

    # Sheet 2: Dues by Department
    ws2 = wb.create_sheet("Dues by Department")
    headers2 = ["Department", "Total Members", "Paid", "Unpaid", "Compliance %"]
    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header(ws2, 1, len(headers2))

    for ri, d in enumerate(dept_summary, 2):
        ws2.cell(row=ri, column=1, value=d["department_name"])
        ws2.cell(row=ri, column=2, value=d["total_members"])
        ws2.cell(row=ri, column=3, value=d["paid_count"])
        ws2.cell(row=ri, column=4, value=d["unpaid_count"])
        ws2.cell(row=ri, column=5, value=f'{d["percentage"]}%')
        pct = d["percentage"]
        green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        yellow_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
        red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        fill = green_fill if pct >= 90 else yellow_fill if pct >= 70 else red_fill
        for ci in range(1, len(headers2) + 1):
            ws2.cell(row=ri, column=ci).fill = fill
    _auto_width(ws2, len(headers2))

    # Sheet 3: Contributions by Department
    ws3 = wb.create_sheet("Contributions by Department")
    headers3 = ["Department", "Post", "Aid Type", "Paid", "Unpaid", "Skipped", "Collection %"]
    for ci, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=ci, value=h)
    _style_header(ws3, 1, len(headers3))

    posts = AidTrackingPost.objects.filter(is_active=True)
    ri = 2
    for post in posts:
        from core_system.services.compliance import contribution_compliance_summary

        dept_contribs = contribution_compliance_summary(post.post_id_PK)
        for dc in dept_contribs:
            ws3.cell(row=ri, column=1, value=dc["department_name"])
            ws3.cell(row=ri, column=2, value=str(post.target_month))
            ws3.cell(row=ri, column=3, value=post.aid_type)
            ws3.cell(row=ri, column=4, value=dc["paid_count"])
            ws3.cell(row=ri, column=5, value=dc["unpaid_count"])
            ws3.cell(row=ri, column=6, value=dc["skipped_count"])
            ws3.cell(row=ri, column=7, value=f'{dc["percentage"]}%')
            ri += 1
    _auto_width(ws3, len(headers3))

    return wb


def generate_department_report(
    dept_id: int,
    year: int | None = None,
    month: int | None = None,
    wb=None,
    sheet_name: str | None = None,
):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    today = timezone.localdate()
    year = year or today.year
    month = month or today.month

    try:
        dept = Department.objects.get(department_id_PK=dept_id)
    except Department.DoesNotExist:
        return None

    members = Member.objects.filter(department_id_FK=dept).exclude(
        membership_status__iexact="retired"
    ).order_by("full_name")

    dues_sheet = sheet_name or "Dues Compliance"
    contrib_sheet = f"{sheet_name} - Contribution Compliance" if sheet_name else "Contribution Compliance"

    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = dues_sheet
    else:
        ws = wb.create_sheet(dues_sheet)
    ws.cell(row=1, column=1, value=f"{dept.name} - Dues Compliance ({year}-{month:02d})").font = Font(
        bold=True, size=12
    )
    ws.merge_cells("A1:E1")

    headers = ["Member Name", "Employee ID", "Dues Status", "Overdue Bucket"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    _style_header(ws, 3, len(headers))

    for ri, m in enumerate(members, 4):
        ws.cell(row=ri, column=1, value=m.full_name)
        ws.cell(row=ri, column=2, value=m.employee_id or "")
        status = member_dues_status(m, year, month)
        ws.cell(row=ri, column=3, value=status)
        ws.cell(row=ri, column=4, value=dues_overdue_bucket(m, year, month) or "")
    _auto_width(ws, len(headers))

    ws2 = wb.create_sheet(contrib_sheet)
    posts = AidTrackingPost.objects.filter(is_active=True)
    headers2 = ["Member Name", "Employee ID", "Post", "Aid Type", "Status"]
    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header(ws2, 1, len(headers2))

    ri = 2
    for m in members:
        for post in posts:
            status = member_contribution_status(m, post.post_id_PK)
            ws2.cell(row=ri, column=1, value=m.full_name)
            ws2.cell(row=ri, column=2, value=m.employee_id or "")
            ws2.cell(row=ri, column=3, value=str(post.target_month))
            ws2.cell(row=ri, column=4, value=post.aid_type)
            ws2.cell(row=ri, column=5, value=status)
            ri += 1
    _auto_width(ws2, len(headers2))

    return wb


def generate_contribution_report(post_id: int | None = None, wb=None):
    from openpyxl import Workbook

    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Contribution Report"
    else:
        ws = wb.create_sheet("Contribution Report")

    posts = AidTrackingPost.objects.filter(is_active=True)
    if post_id:
        posts = posts.filter(post_id_PK=post_id)

    headers = [
        "Post ID", "Aid Type", "Target Month", "Department",
        "Member Name", "Employee ID", "Status", "Expected Amount",
        "Paid Amount", "Payment Date",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    _style_header(ws, 1, len(headers))

    ri = 2
    for post in posts:
        contribs = Contribution.objects.filter(
            aid_tracking_post_id_FK=post,
        ).select_related("member_id_FK__department_id_FK")
        for c in contribs:
            dept_name = (
                c.member_id_FK.department_id_FK.name
                if c.member_id_FK.department_id_FK
                else "N/A"
            )
            ws.cell(row=ri, column=1, value=post.post_id_PK)
            ws.cell(row=ri, column=2, value=post.aid_type)
            ws.cell(row=ri, column=3, value=post.target_month)
            ws.cell(row=ri, column=4, value=dept_name)
            ws.cell(row=ri, column=5, value=c.member_id_FK.full_name)
            ws.cell(row=ri, column=6, value=c.member_id_FK.employee_id or "")
            ws.cell(row=ri, column=7, value=c.status)
            ws.cell(row=ri, column=8, value=_fmt(c.expected_amount))
            ws.cell(row=ri, column=9, value=_fmt(c.paid_amount))
            ws.cell(row=ri, column=10, value=str(c.payment_date) if c.payment_date else "")
            ri += 1
    _auto_width(ws, len(headers))

    return wb


def approve_auditor_report(report_id: int, officer, remarks: str = "") -> dict | None:
    from core_system.models import AuditFindingsReport

    try:
        report = AuditFindingsReport.objects.get(
            audit_report_id_PK=report_id,
            report_status="Submitted",
        )
    except AuditFindingsReport.DoesNotExist:
        return None

    report.report_status = "Approved"
    report.certification_status = "Certified"
    report.certified_by_user_id_FK = officer
    report.save(update_fields=["report_status", "certification_status", "certified_by_user_id_FK"])

    return {
        "report_id": report.audit_report_id_PK,
        "status": report.report_status,
        "certification_status": report.certification_status,
    }


def request_report_revision(report_id: int, officer, remarks: str = "") -> dict | None:
    from core_system.models import AuditFindingsReport

    try:
        report = AuditFindingsReport.objects.get(
            audit_report_id_PK=report_id,
            report_status="Submitted",
        )
    except AuditFindingsReport.DoesNotExist:
        return None

    report.report_status = "Revision Requested"
    report.findings_summary += f"\n\n--- Revision requested by {officer.full_name}: {remarks}"
    report.save(update_fields=["report_status", "findings_summary"])

    return {
        "report_id": report.audit_report_id_PK,
        "status": report.report_status,
    }


def create_auditor_report(year: int, month: int, officer) -> dict:
    from core_system.models import AuditFindingsReport

    dept_summary = dues_compliance_summary(year, month)
    total_members = sum(d["total_members"] for d in dept_summary)
    total_paid = sum(d["paid_count"] for d in dept_summary)
    total_unpaid = sum(d["unpaid_count"] for d in dept_summary)
    pct = round(total_paid / total_members * 100, 1) if total_members else 0.0

    period = f"{year}-{month:02d}"
    low_depts = [d for d in dept_summary if d["percentage"] < 70]

    findings = (
        f"Compliance report for period {period}.\n"
        f"Overall compliance: {pct}%\n"
        f"Total active members: {total_members}\n"
        f"Paid: {total_paid}, Unpaid: {total_unpaid}\n"
    )
    if low_depts:
        findings += f"Departments below 70% threshold: {', '.join(d['department_name'] for d in low_depts)}\n"

    report = AuditFindingsReport.objects.create(
        report_title=f"Compliance Report - {period}",
        report_period=period,
        findings_summary=findings,
        report_status="Draft",
        prepared_by_user_id_FK=officer,
        prepared_date=timezone.localdate(),
        presentation_status="Draft",
        certification_status="Pending",
    )

    return {
        "report_id": report.audit_report_id_PK,
        "title": report.report_title,
        "period": report.report_period,
        "findings_summary": report.findings_summary,
        "status": report.report_status,
        "prepared_date": str(report.prepared_date),
    }


def generate_organization_fund_report(year: int, month: int, report_type: str = "monthly", wb=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from datetime import date
    from calendar import monthrange

    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Fund Summary"
    else:
        ws = wb.create_sheet("Fund Summary")

    period_start = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    period_end = date(year, month, last_day)
    period_label = f"{year}-{month:02d}"

    start_dt = pytz.UTC.localize(datetime.combine(period_start, datetime.min.time()))
    end_dt = pytz.UTC.localize(datetime.combine(period_end, datetime.max.time().replace(microsecond=0)))

    transactions = FundTransaction.objects.filter(
        recorded_at__gte=start_dt,
        recorded_at__lte=end_dt,
    ).select_related("recorded_by_user_id_FK").order_by("recorded_at")

    inflows = [t for t in transactions if t.direction == "inflow"]
    outflows = [t for t in transactions if t.direction == "outflow"]

    total_inflow = sum(float(t.amount) for t in inflows)
    total_outflow = sum(float(t.amount) for t in outflows)
    net = total_inflow - total_outflow

    contributions = Contribution.objects.filter(
        payment_date__gte=period_start,
        payment_date__lte=period_end,
        status__in=["PAID", "RECORDED", "PENDING_VERIFICATION"],
    ).select_related("member_id_FK", "aid_tracking_post_id_FK").order_by("payment_date")

    total_contributions = sum(float(c.paid_amount) for c in contributions)

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=f"Organization Fund Report - {period_label}").font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    headers = ["Metric", "Value"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    _style_header(ws, 3, 2)

    data_rows = [
        ("Period", period_label),
        ("Report Type", "Weekly" if report_type == "weekly" else "Monthly"),
        ("Total Inflows", total_inflow),
        ("Total Outflows", total_outflow),
        ("Net Fund Position", net),
        ("Total Contributions Collected", total_contributions),
        ("Number of Transactions", len(transactions)),
    ]
    for ri, (metric, val) in enumerate(data_rows, 4):
        ws.cell(row=ri, column=1, value=metric)
        ws.cell(row=ri, column=2, value=val)
    _auto_width(ws, 2)

    # Sheet 2: Inflows
    ws2 = wb.create_sheet("Inflows")
    headers2 = ["Date", "Source Type", "Source ID", "Amount", "Description", "Recorded By"]
    for ci, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header(ws2, 1, len(headers2))
    for ri, t in enumerate(inflows, 2):
        ws2.cell(row=ri, column=1, value=str(t.recorded_at.date()) if t.recorded_at else "")
        ws2.cell(row=ri, column=2, value=t.source_type or "")
        ws2.cell(row=ri, column=3, value=t.source_id or "")
        ws2.cell(row=ri, column=4, value=float(t.amount))
        ws2.cell(row=ri, column=5, value=t.description or "")
        ws2.cell(row=ri, column=6, value=t.recorded_by_user_id_FK.full_name if t.recorded_by_user_id_FK else "")
    _auto_width(ws2, len(headers2))

    # Sheet 3: Outflows
    ws3 = wb.create_sheet("Outflows")
    headers3 = ["Date", "Source Type", "Source ID", "Amount", "Description", "Recorded By"]
    for ci, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=ci, value=h)
    _style_header(ws3, 1, len(headers3))
    for ri, t in enumerate(outflows, 2):
        ws3.cell(row=ri, column=1, value=str(t.recorded_at.date()) if t.recorded_at else "")
        ws3.cell(row=ri, column=2, value=t.source_type or "")
        ws3.cell(row=ri, column=3, value=t.source_id or "")
        ws3.cell(row=ri, column=4, value=float(t.amount))
        ws3.cell(row=ri, column=5, value=t.description or "")
        ws3.cell(row=ri, column=6, value=t.recorded_by_user_id_FK.full_name if t.recorded_by_user_id_FK else "")
    _auto_width(ws3, len(headers3))

    # Sheet 4: Contributions
    ws4 = wb.create_sheet("Contributions")
    headers4 = ["Date", "Member Name", "Employee ID", "Department", "Post ID", "Aid Type", "Expected", "Paid", "Status"]
    for ci, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=ci, value=h)
    _style_header(ws4, 1, len(headers4))
    for ri, c in enumerate(contributions, 2):
        member = c.member_id_FK
        post = c.aid_tracking_post_id_FK
        ws4.cell(row=ri, column=1, value=str(c.payment_date) if c.payment_date else "")
        ws4.cell(row=ri, column=2, value=member.full_name if member else "")
        ws4.cell(row=ri, column=3, value=member.employee_id or "" if member else "")
        ws4.cell(row=ri, column=4, value=member.department or "" if member else "")
        ws4.cell(row=ri, column=5, value=post.post_id_PK if post else "")
        ws4.cell(row=ri, column=6, value=post.aid_type if post else "")
        ws4.cell(row=ri, column=7, value=float(c.expected_amount))
        ws4.cell(row=ri, column=8, value=float(c.paid_amount))
        ws4.cell(row=ri, column=9, value=c.status)
    _auto_width(ws4, len(headers4))

    return wb


def create_organization_fund_report(officer, year: int, month: int, report_type: str = "monthly") -> dict:
    period = f"{year}-{month:02d}"

    wb = generate_organization_fund_report(year, month, report_type)

    filename = f"fund_report_{report_type}_{period}.xlsx"
    file_path = f"reports/{filename}"
    default_storage_path = getattr(settings, "MEDIA_ROOT", "")
    full_path = os.path.join(str(default_storage_path), file_path) if default_storage_path else file_path

    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    wb.save(full_path)

    report = OrganizationFundReport.objects.create(
        report_period=period,
        report_type=report_type,
        report_status="Draft",
        file_path=file_path,
        prepared_by_user_id_FK=officer,
    )

    return {
        "report_id": report.report_id_PK,
        "period": report.report_period,
        "report_type": report.report_type,
        "status": report.report_status,
        "file_path": report.file_path,
    }


def generate_unified_report(
    year: int | None = None,
    month: int | None = None,
    sections: list[str] | None = None,
) -> Workbook:
    from openpyxl import Workbook

    today = timezone.localdate()
    year = year or today.year
    month = month or today.month

    if sections is None:
        sections = ["overall", "department", "contribution", "fund"]

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    if "overall" in sections:
        generate_overall_report(year, month, wb)

    if "department" in sections:
        departments = Department.objects.filter(is_active=True).order_by("name")
        for dept in departments:
            generate_department_report(dept.department_id_PK, year, month, wb, sheet_name=dept.name)

    if "contribution" in sections:
        generate_contribution_report(None, wb)

    if "fund" in sections:
        generate_organization_fund_report(year, month, "monthly", wb)

    return wb
